from datetime import date, datetime
from io import BytesIO
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session as DBSession
from app.database import get_db
from app.models.caces_obtenu import CacesObtenu
from app.models.stagiaire import Stagiaire
from app.models.session import Session as SessionModel

router = APIRouter(prefix="/api/registre-caces", tags=["Registre CACES"])


def _nature(co: CacesObtenu) -> str:
    """3 natures : OTC (interne, y compris repris), sous-traitance, externe."""
    if not co.organisme_externe:
        return "otc"
    return "st" if co.sous_traitance else "ext"


def _mois_entre(a: date, b: date) -> int:
    return (b.year - a.year) * 12 + (b.month - a.month)


def _statut_echeance(ech: date, aujourdhui: date, seuil_mois: int) -> str:
    if ech is None:
        return "val"
    if ech < aujourdhui:
        return "exp"
    if _mois_entre(aujourdhui, ech) <= seuil_mois:
        return "ren"
    return "val"


@router.get("")
def registre_caces(seuil: int = 6, db: DBSession = Depends(get_db)):
    """Vue a plat de tous les CACES obtenus (relance / complement).
    Exclut les CACES annules. Le calcul du statut d'echeance se fait cote
    serveur avec le seuil recu (mois). Filtres/tri restants cote front.
    """
    aujourdhui = date.today()

    # Maps stagiaire + session (evite les N+1)
    stagiaires = {s.id: s for s in db.query(Stagiaire).all()}
    sessions = {s.id: s for s in db.query(SessionModel).all()}

    records = (
        db.query(CacesObtenu)
        .filter(CacesObtenu.statut != "annule")
        .all()
    )

    lignes = []
    for co in records:
        s = stagiaires.get(co.stagiaire_id)
        sess = sessions.get(co.session_id)
        ech = co.date_echeance
        lignes.append({
            "id": co.id,
            "stagiaire_id": co.stagiaire_id,
            "nom": s.nom if s else "?",
            "prenom": s.prenom if s else "?",
            "societe": (s.employeur if s else "") or "",
            "famille": co.famille,
            "categorie": co.categorie,
            "options_obtenues": co.options_obtenues or "",
            "nature": _nature(co),
            "numero": co.ancien_numero or (str(co.numero_ordre).zfill(4) if co.numero_ordre else ""),
            "date_obtention": co.date_obtention.isoformat() if co.date_obtention else None,
            "date_echeance": ech.isoformat() if ech else None,
            "statut_echeance": _statut_echeance(ech, aujourdhui, seuil),
            "session_reference": (sess.reference if sess and sess.reference else (f"Session {co.session_id}" if co.session_id else "")),
        })

    # Tri par defaut : echeance croissante (les plus urgents en tete), None en fin
    lignes.sort(key=lambda x: (x["date_echeance"] is None, x["date_echeance"] or ""))

    # Listes distinctes pour peupler les menus deroulants cote front
    societes = sorted({l["societe"] for l in lignes if l["societe"]})
    familles = sorted({l["famille"] for l in lignes if l["famille"]})

    return {
        "seuil": seuil,
        "aujourdhui": aujourdhui.isoformat(),
        "total": len(lignes),
        "societes": societes,
        "familles": familles,
        "lignes": lignes,
    }


@router.get("/export")
def registre_caces_export(
    seuil: int = 6,
    soc: str = "",
    fam: str = "",
    nat: str = "",
    txt: str = "",
    exp: str = "1",
    ren: str = "1",
    val: str = "0",
    db: DBSession = Depends(get_db),
):
    """Export Excel de la vue filtree (memes filtres que le front)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    _LIB_NAT = {"otc": "Organisme (OTC)", "st": "Sous-traitance", "ext": "Externe"}
    _LIB_STA = {"exp": "Expire", "ren": "A renouveler", "val": "Valide"}
    _show = {"exp": exp == "1", "ren": ren == "1", "val": val == "1"}
    _txt = (txt or "").lower()

    data = registre_caces(seuil=seuil, db=db)
    lignes = data["lignes"]

    def _garde(l):
        if soc and l["societe"] != soc:
            return False
        if fam and l["famille"] != fam:
            return False
        if nat and l["nature"] != nat:
            return False
        if _txt and _txt not in (l["nom"] + " " + l["prenom"] + " " + l["societe"]).lower():
            return False
        if not _show.get(l["statut_echeance"], True):
            return False
        return True

    lignes = [l for l in lignes if _garde(l)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registre CACES"

    entetes = ["Nom", "Prenom", "Societe", "Famille", "Categorie", "Options",
               "Nature", "N", "Obtention", "Echeance", "Statut", "Session"]
    ws.append(entetes)

    head_fill = PatternFill("solid", fgColor="2D2D2D")
    head_font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="left", vertical="center")

    def _fr(iso):
        if not iso:
            return ""
        p = iso.split("-")
        return p[2] + "/" + p[1] + "/" + p[0]

    couleur_sta = {"exp": "A32D2D", "ren": "854F0B", "val": "3B6D11"}
    for l in lignes:
        ws.append([
            l["nom"], l["prenom"], l["societe"], l["famille"], l["categorie"],
            l["options_obtenues"], _LIB_NAT.get(l["nature"], l["nature"]),
            l["numero"], _fr(l["date_obtention"]), _fr(l["date_echeance"]),
            _LIB_STA.get(l["statut_echeance"], l["statut_echeance"]),
            l["session_reference"],
        ])
        cell_sta = ws.cell(row=ws.max_row, column=11)
        cell_sta.font = Font(color=couleur_sta.get(l["statut_echeance"], "000000"), bold=True)

    largeurs = [16, 14, 22, 9, 9, 18, 18, 10, 12, 12, 15, 16]
    for i, w in enumerate(largeurs, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    ws.append([])
    stamp = datetime.now().strftime("%d/%m/%Y %H:%M")
    filtres = []
    if soc:
        filtres.append("Societe=" + soc)
    if fam:
        filtres.append("Famille=" + fam)
    if nat:
        filtres.append("Nature=" + _LIB_NAT.get(nat, nat))
    if txt:
        filtres.append("Recherche=" + txt)
    etats = [k for k, v in _show.items() if v]
    filtres.append("Echeance=" + ",".join(_LIB_STA.get(e, e) for e in etats))
    filtres.append("Seuil=" + str(seuil) + " mois")
    ws.append(["Export du " + stamp + "  |  " + str(len(lignes)) + " ligne(s)  |  " + "  ".join(filtres)])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="6B6A65", size=9)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nom = "registre_caces_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="' + nom + '"'},
    )
